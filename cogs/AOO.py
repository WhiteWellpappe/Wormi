import discord, datetime
from discord.ext import commands
from openpyxl import Workbook, load_workbook

def setup(bot):
    bot.add_cog(AOO(bot))


class AOO(commands.Cog, name='AOO'):
    def __init__(self, bot):
        self.bot = bot

    #Resetting AOO Sheet
    @commands.command(name="resetaoo", help="admin only, clears aoo.xlsx")
    @commands.has_any_role('Discord King', 'Leader')
    async def resetaoo(self, ctx):
        wb1=load_workbook('./aoo_default.xlsx')
        wb1.save("./aoo.xlsx")
        await ctx.send(f"{ctx.author.mention} AOO Sheet got reseted.")
    
    #Cleaning the Aoo-registration-channel
    @commands.command(name="cleanaoo", help="admin only, cleans 200 #aoo-registration messages")
    @commands.has_any_role('Discord King', 'Leader')
    async def cleanaoo(self, ctx):
        clean=self.bot.get_channel(651033088943587328)
        await clean.purge(limit=200)
        await ctx.send(f"{ctx.author.mention} {clean.mention} cleaned.")

    #Sending the current aoo.xlsx
    @commands.command(name="sendaoo", help="sends current aoo.xlsx")
    @commands.has_any_role('Discord King', 'Leader', 'Officer')
    async def sendaoo(self, ctx):
        path='./aoo.xlsx'
        await ctx.send(f'{ctx.author.mention}')
        await ctx.send(file=discord.File(path))

    #Members of largest AOO roster, automatically selects Saturday or Sunday
    @commands.command(name="aoolist", help="mentions all members of the largest WgD aoo list")
    @commands.has_any_role('Discord King', 'Officer', 'Leader')
    async def aoolist(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        cell_a=ws.cell(row=2, column=11)
        cell_b=ws.cell(row=3, column=11)
        liste=""
        try:
            a=int(cell_a.value)
            b=int(cell_b.value)
        except:
            pass
        if a==b:
            await ctx.send(f"{ctx.author.mention} Error. Both groups are the same size. Total Members: {a}")
        elif a>b:
            await ctx.send(f"{ctx.author.mention} Participation List for **SATURAY 20UTC**. Total Members: {a}")
            for s in range(3, 34):
                sc=ws.cell(row=s, column=3)
                if sc.value==None:
                    break
                else:
                    liste=liste+str(sc.value)+" "
        elif a<b:
            await ctx.send(f"{ctx.author.mention} Participation List for **SUNDAY 20UTC**. Total Members: {b}")
            for s in range(3, 34):
                sc=ws.cell(row=s, column=6)
                if sc.value==None:
                    break
                else:
                    liste=liste+str(sc.value)+" "
        if liste!="":
            await ctx.send(liste)

    #Members WgD Saturday roster
    @commands.command(name="aoolistsat", help="mentions all members of the aoo list for saturday")
    @commands.has_any_role('Discord King', 'Officer', 'Leader')
    async def aoolistsat(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        liste=""
        cell_a=ws.cell(row=2, column=11)
        try:
            a=int(cell_a.value)
        except:
            pass
        for s in range(3, 34):
            sc=ws.cell(row=s, column=3)
            if sc.value==None:
                break
            else:
                liste=liste+str(sc.value)+" "
        await ctx.send(f"{ctx.author.mention} Participation List for **SATURAY 20UTC**. Total Members: {a}")
        if liste!="":
            await ctx.send(liste)


    #Members WgD Sunday roster
    @commands.command(name="aoolistsun", help="mentions all members of the aoo list for sunday")
    @commands.has_any_role('Discord King', 'Officer', 'Leader')
    async def aoolistsun(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        liste=""
        cell_b=ws.cell(row=3, column=11)
        try:
            b=int(cell_b.value)
        except:
            pass
        for s in range(3, 34):
            sc=ws.cell(row=s, column=6)
            if sc.value==None:
                break
            else:
                liste=liste+str(sc.value)+" "
        await ctx.send(f"{ctx.author.mention} Participation List for **SUNDAY 20UTC**. Total Members: {b}")
        if liste!="":
            await ctx.send(liste)


    #Members WgDF roster
    @commands.command(name="aoolistF", help="mentions all members of the WgDF aoo list")
    @commands.has_any_role('Discord King', 'Officer', 'Leader')
    async def aoolistF(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        cell_c=ws.cell(row=4, column=11)
        liste=""
        try:
            c=int(cell_c.value)
        except:
            pass
        await ctx.send(f"{ctx.author.mention} Participation List for **WgDF Sunday 4UTC**. Total Members: {c}")
        for s in range(3, 34):
            sc=ws.cell(row=s, column=9)
            if sc.value==None:
                break
            else:
                liste=liste+str(sc.value)+" "
        if liste!="":
            await ctx.send(liste)


    #Membercount all 3 rosters
    @commands.command(name="aoomember", help="gives out the member count of all aoo lists")
    @commands.has_any_role('Discord King', 'Officer', 'Leader')
    async def aoomember(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        cella=ws.cell(row=2, column=11)
        cellb=ws.cell(row=3, column=11)
        cellc=ws.cell(row=4, column=11)
        try:
            a=int(cella.value)
            b=int(cellb.value)
            c=int(cellc.value)
        except:
            pass
        await ctx.send(f"{ctx.author.mention} Members Saturday: {a}, Members Sunday: {b}, Members Farm: {c}")


    #Votes for both days
    @commands.command(name="aoovotes", help="shows currents AOO votes")
    @commands.has_any_role('Discord King', 'Officer', 'Leader')
    async def aoovotes(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        sat=ws.cell(row=5, column=11).value
        sun=ws.cell(row=6, column=11).value
        await ctx.send(f"{ctx.author.mention} Votes for Saturday: {sat} Votes for Sunday: {sun}")


    @commands.command(name="leaveaoo", help="delete your registration and void your vote in the current aoo voting")
    @commands.has_any_role("Discord King", "Officer", "Leader", "Member")
    async def leaveaoo(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        mention=ctx.author.mention
        for i in range(3, 33):
            c=ws.cell(row=i, column=3)
            if c.value!=mention:
                continue
            else:
                ws.cell(row=i, column=1, value="")
                ws.cell(row=i, column=2, value="")
                ws.cell(row=i, column=3, value="")
                sat=ws.cell(row=2, column=11).value-1
                ws.cell(row=2, column=11, value=sat)
                break
        for n in range(3, 33):
            d=ws.cell(row=n, column=6)
            if d.value!=mention:
                continue
            else:
                ws.cell(row=n, column=4, value="")
                ws.cell(row=n, column=5, value="")
                ws.cell(row=n, column=6, value="")
                sun=ws.cell(row=3, column=11).value-1
                ws.cell(row=3, column=11, value=sun)
                break
        for m in range(3, 33):
            c=ws.cell(row=m, column=9)
            if c.value!=mention:
                continue
            else:
                ws.cell(row=m, column=7, value="")
                ws.cell(row=m, column=8, value="")
                ws.cell(row=m, column=9, value="")
                farm=ws.cell(row=4, column=11).value-1
                ws.cell(row=4, column=11, value=farm)
                break
        for k in range(1, 33):
            c=ws.cell(row=k, column=12)
            if c.value!=mention:
                continue
            else:
                ws.cell(row=k, column=12, value="")
                sat=ws.cell(row=5, column=11).value-1
                ws.cell(row=5, column=11, value=sat)
                break
        for q in range(1, 33):
            c=ws.cell(row=q, column=13)
            if c.value!=mention:
                continue
            else:
                ws.cell(row=q, column=13, value="")
                sat=ws.cell(row=6, column=11).value-1
                ws.cell(row=6, column=11, value=sat)
                break
        wb.save('./aoo.xlsx') 
        await ctx.send(f"{ctx.author.mention} Your registration and vote has been voided.")


    @commands.command(name="kickaoo", help="deletes registration and vote of the mentioned member")
    @commands.has_any_role("Discord King", "Officer", "Leader")
    async def kickaoo(self, ctx, mention):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        for i in range(3, 33):
            c=ws.cell(row=i, column=3)
            if c.value!=mention:
                continue
            else:
                ws.cell(row=i, column=1, value="")
                ws.cell(row=i, column=2, value="")
                ws.cell(row=i, column=3, value="")
                sat=ws.cell(row=2, column=11).value-1
                ws.cell(row=2, column=11, value=sat)
                break
        for n in range(3, 33):
            d=ws.cell(row=n, column=6)
            if d.value!=mention:
                continue
            else:
                ws.cell(row=n, column=4, value="")
                ws.cell(row=n, column=5, value="")
                ws.cell(row=n, column=6, value="")
                sun=ws.cell(row=3, column=11).value-1
                ws.cell(row=3, column=11, value=sun)
                break
        for m in range(3, 33):
            c=ws.cell(row=m, column=9)
            if c.value!=mention:
                continue
            else:
                ws.cell(row=m, column=7, value="")
                ws.cell(row=m, column=8, value="")
                ws.cell(row=m, column=9, value="")
                farm=ws.cell(row=4, column=11).value-1
                ws.cell(row=4, column=11, value=farm)
                break
        for k in range(1, 33):
            c=ws.cell(row=k, column=12)
            if c.value!=mention:
                continue
            else:
                ws.cell(row=k, column=12, value="")
                sat=ws.cell(row=5, column=11).value-1
                ws.cell(row=5, column=11, value=sat)
                break
        for q in range(1, 33):
            c=ws.cell(row=q, column=13)
            if c.value!=mention:
                continue
            else:
                ws.cell(row=q, column=13, value="")
                sat=ws.cell(row=6, column=11).value-1
                ws.cell(row=6, column=11, value=sat)
                break
        wb.save('./aoo.xlsx') 
        await ctx.send(f"{ctx.author.mention} {mention} has been kicked from the aoo roster.")



    #AOO announcement post
    @commands.command(name="aoo", help="aoo announcement, starting on saturday, the Xth")
    @commands.has_any_role('Discord King', 'Officer', 'Leader')
    async def aoo (self, ctx, arg):
        if arg==None:
            ctx.send(f"{ctx.author.mention} Dont forget the date! ;)")
        else:
            try:
                _=int(arg)
            except:
                await ctx.send(f"{ctx.author.mention} Please check the command again. An error occoured and I caught it so you don't wreck the channel like Woody did once.")
            arg2=str(int(arg)+1)
            #sorting months into 2 groups
            m30=["4", "6", "9", "11"]
            m31=["1", "3", "5", "7", "8", "10", "12"]
            #checking changing month mid-event
            if int(arg)>31 or int(arg)<=0:
                await ctx.send("Please insert a valid day.")
            else:
                if arg=="28":
                    if datetime.datetime.now().strftime("%m")=="02":
                        arg2="1" if int(datetime.datetime.now().strftime("%Y"))%4!=0 else None
                elif arg=="29":
                    if datetime.datetime.now().strftime("%m")=="02":
                        arg2="1" if int(datetime.datetime.now().strftime("%Y"))%4==0 else None                            
                elif arg=="30":
                    for x in m30:
                        arg2="1" if x==datetime.datetime.now().strftime("%m") else None
                elif arg=="31":
                    for x in m31:
                        arg2='1' if x==datetime.datetime.now().strftime("%m") else None
                #correcting the ending
                if int(arg)==1:
                    arg=arg+"st"
                elif int(arg)==2:
                    arg=arg+"nd"
                elif int(arg)==3:
                    arg=arg+"rd"
                else:
                    arg=arg+"th"   
                if int(arg2)==1:
                    arg2=arg2+"st"
                elif int(arg2)==2:
                    arg2=arg2+"nd"
                elif int(arg2)==3:
                    arg2=arg2+"rd"
                else:
                    arg2=arg2+"th"
                await ctx.send(f"""@everyone

:loudspeaker: AOO Voting!

FOR WgD (Main alliance)
1 - Only Sat, {arg}, 20:00 UTC
2 - Only Sun, {arg2}, 20:00 UTC
3a - Can do both, vote for/prefere saturday
3b - Can do both, vote for/prefere sunday

FOR WgDF (Farm alliance)
F - Sun, {arg2}, 04:00 UTC

:pizza: Write !joinaoo followed by your favourite option, send through comment box
        **Example: !joinaoo 3a**
:rice: On successfull enlisting Wormi will give a feedback.
:taco: **25 slots, 5 reserved**, First Come First Serve system
:hamburger: Start voting now, end on Thursday or when full
:pie: You can void all registrations and your vote with just **!leaveaoo**
:candy: You can alter your vote ONLY WITH 3a and 3b. If you need to change the single participation day, use !leaveaoo and register again

:fries: If you vote for WgD, DO NOT duplicate your vote in WgDF unless you want to participate with your farm account
:cheese: If someone enlists for both Main and Farm we assume your farm will play AOO in WgDF

:carrot: If you encounter an error, please hit up any R4 right away.""")
        if ctx.channel.id==651033088943587328:
            #enabling writing/reading/history for members
            role=ctx.guild.get_role(530162542148976661)
            await ctx.channel.set_permissions(role, read_messages=True, send_messages=True, view_channel=True, read_message_history=True)



#Functions for !joinaoo_____________________________________________________________________________________________________________________________________________________________    
    def aoosat_full(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        global space
        space=0
        n=ws.cell(row=27, column=1).value
        if n!=None:
            space=1
        return space


    def aoosat_check(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        mention=ctx.author.mention
        global registered
        registered=0
        for m in range(3, 33):
            n=ws.cell(row=m, column=3)
            if n.value==mention:
                registered=1
                break
        return registered

    def aoosat_reg(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        author=ctx.author.display_name
        mention=ctx.author.mention
        for i in range(3, 33):
            c=ws.cell(row=i, column=1)
            if c.value!=None:
                continue
            else:
                ws.cell(row=i, column=1, value=i-2)
                ws.cell(row=i, column=2, value=author)
                ws.cell(row=i, column=3, value=mention)
                sat=ws.cell(row=2, column=11).value+1
                ws.cell(row=2, column=11, value=sat)
                break  
        wb.save('./aoo.xlsx') 

    def aoosun_full(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        global space
        space=0
        n=ws.cell(row=27, column=4).value
        if n!=None:
            space=1
        return space

    def aoosun_check(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        mention=ctx.author.mention
        global registered
        registered=0
        for m in range(3, 33):
            n=ws.cell(row=m, column=6)
            if n.value==mention:
                registered=1
                break
        return registered

    def aoosun_reg(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        author=ctx.author.display_name
        mention=ctx.author.mention
        for i in range(3, 33):
            c=ws.cell(row=i, column=4)
            if c.value!=None:
                continue
            else:
                ws.cell(row=i, column=4, value=i-2)
                ws.cell(row=i, column=5, value=author)
                ws.cell(row=i, column=6, value=mention)
                sun=ws.cell(row=3, column=11).value+1
                ws.cell(row=3, column=11, value=sun)
                break
        wb.save('./aoo.xlsx') 

    def aooF(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        author=ctx.author.display_name
        mention=ctx.author.mention
        global registered
        #check if already registered
        for m in range(3, 34):
            n=ws.cell(row=m, column=9)
            if n.value==mention:
                registered=1
                return registered
        #if not registered, insert data into next empty column
        if registered==0:
            for i in range(3, 34):
                c=ws.cell(row=i, column=7)
                if c.value!=None:
                    continue
                else:
                    ws.cell(row=i, column=7, value=i-2)
                    ws.cell(row=i, column=8, value=author)
                    ws.cell(row=i, column=9, value=mention)
                    ws.cell(row=4, column=11, value=i-2)
                    break   
        wb.save('./aoo.xlsx')

    def sat_checkvote(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        mention=ctx.author.mention
        vote=0
        for i in range(1, 33):
            n=ws.cell(row=i, column=12)
            if n.value==mention:
                vote=1
                break
        return vote
    
    def sun_checkvote(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        mention=ctx.author.mention
        vote=0
        for i in range(1, 33):
            n=ws.cell(row=i, column=13)
            if n.value==mention:
                vote=1
                break
        return vote

    def sat_addvote(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        sat=ws.cell(row=5, column=11).value+1
        ws.cell(row=5, column=11, value=sat)
        mention=ctx.author.mention
        for i in range(2, 33):
            c=ws.cell(row=i, column=12)
            if c.value!=None:
                continue
            else:
                ws.cell(row=i, column=12, value=mention)
                break
        wb.save('./aoo.xlsx')

    def sun_addvote(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        sun=ws.cell(row=6, column=11).value+1
        ws.cell(row=6, column=11, value=sun)
        mention=ctx.author.mention
        for i in range(2, 33):
            c=ws.cell(row=i, column=13)
            if c.value!=None:
                continue
            else:
                ws.cell(row=i, column=13, value=mention)
                break
        wb.save('./aoo.xlsx')

    def sat_deletevote(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        mention=ctx.author.mention
        for i in range(1, 33):
            c=ws.cell(row=i, column=12)
            if c.value==mention:
                sat=ws.cell(row=5, column=11).value-1
                ws.cell(row=5, column=11, value=sat)
                z=""
                ws.cell(row=i, column=12, value=z)
                break
            else:
                continue
        wb.save('./aoo.xlsx')

    def sun_deletevote(self, ctx):
        wb=load_workbook('./aoo.xlsx')
        ws=wb.active
        mention=ctx.author.mention
        for i in range(1, 33):
            c=ws.cell(row=i, column=13)
            if c.value==mention:
                sun=ws.cell(row=6, column=11).value-1
                ws.cell(row=6, column=11, value=sun)
                z=""
                ws.cell(row=i, column=13, value=z)             
                break
            else:
                continue
        wb.save('./aoo.xlsx')


    #Adding people to AOO Sheet with Discord_________________________________________________________________________________________________________________________________________
    @commands.command(name="joinaoo", help="command to select the desired option to join the AOO team")
    @commands.has_any_role('Discord King', 'Officier', 'Leader', 'Member')
    async def joinaoo(self, ctx, arg):
        global registered
        registered=0
        regsat=1
        regsun=1
        if ctx.channel.id==651033088943587328:
            #registering for saturday
            if arg=="1":
                if self.aoosat_check(ctx)==0:
                    if self.aoosat_full(ctx)==0:
                        self.aoosat_reg(ctx)
                        self.sat_addvote(ctx)
                        if self.sun_checkvote(ctx)==1:
                            self.sun_deletevote(ctx)
                            await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SATURDAY 20 UTC** You changed your vote to Saturday.")
                        else:
                            await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SATURDAY 20 UTC** You votet for the same day.")
                    else:
                        await ctx.send(f"{ctx.author.mention} Sadly the roster for SATURDAY 20 UTC is already full. Please consider playing in WgDF or CIMA.")
                else:
                    await ctx.send(f"{ctx.author.mention} is already registered for Saturday 20 UTC.")
            #registering for sunday
            elif arg=="2":
                if self.aoosun_check(ctx)==0:
                    if self.aoosun_full(ctx)==0:
                        self.aoosun_reg(ctx)
                        self.sun_addvote(ctx)
                        if self.sat_checkvote==1:
                            self.sat_deletevote(ctx)
                            await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SUNDAY 20 UTC** You changed your vote to Sunday.")
                        else:
                            await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SUNDAY 20 UTC** You votet for the same day.")
                    else:
                        await ctx.send(f"{ctx.author.mention} Sadly the roster for SUNDAY 20 UTC is already full. Please consider playing in WgDF or CIMA.")
                else:
                    await ctx.send(f"{ctx.author.mention} is already registered for Sunday 20 UTC.")
            #registering for both, prefering sat
            elif arg=="3a":
                if self.sat_checkvote(ctx)==0:
                    self.sat_addvote(ctx)   
                if self.sun_checkvote(ctx)==1:
                    self.sun_deletevote(ctx)
                if self.aoosat_check(ctx)==0:
                    regsat=0
                if self.aoosun_check(ctx)==0:
                    regsun=0
                #free for both
                if regsat==0 and regsun==0:
                    if self.aoosat_full(ctx)==0:
                        self.aoosat_reg(ctx)
                        if self.aoosun_full(ctx)==0:
                            self.aoosun_reg(ctx)
                            await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SATURDAY 20 UTC** and **SUNDAY 20 UTC**. You voted for Saturday.")
                        else:
                            await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SATURDAY 20 UTC**. Sunday is already full. You voted for Saturday.")
                    else:
                        if self.aoosun_full(ctx)==0:
                            self.aoosun_reg(ctx)
                            self.sat_deletevote(ctx)
                            self.sun_addvote(ctx)
                            await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SUNDAY 20 UTC**. Saturday is already full. You voted for Sunday instead.")
                        else:
                            self.sat_deletevote(ctx)
                            await ctx.send(f"{ctx.author.mention} Sadly both days are full. You didn't vote.")
                #already registered for sun
                elif regsat==0 and regsun==1:
                    if self.aoosat_full(ctx)==0:
                        self.aoosat_reg(ctx)
                        await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SATURDAY 20 UTC**. You were already registered for Sunday. You changed your vote to Saturday.")
                    else:
                        self.sat_deletevote(ctx)
                        self.sun_addvote(ctx) 
                        await ctx.send(f"{ctx.author.mention} Saturday is full already. You were already registered for Sunday. Your vote didn't change.")
                #already registered for sat
                elif regsat==1 and regsun==0:
                    if self.aoosun_full(ctx)==0:
                        self.aoosun_reg(ctx)
                        await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SUNDAY 20 UTC**. You were already registered for Saturday. Your vote didn't change.")
                    else:
                        await ctx.send(f"{ctx.author.mention} Sunday is full already. Nothing changed.")
                #already registered for both
                elif regsat==1 and regsun==1:
                    await ctx.send(f" {ctx.author.mention} You are already registered for both days. Your voted for Saturday")
            #registering for both, prefering sun
            elif arg=="3b":
                if self.sat_checkvote(ctx)==1:
                    self.sat_deletevote(ctx)   
                if self.sun_checkvote(ctx)==0:
                    self.sun_addvote(ctx)
                if self.aoosat_check(ctx)==0:
                    regsat=0
                if self.aoosun_check(ctx)==0:
                    regsun=0
                #free for both
                if regsat==0 and regsun==0:
                    if self.aoosat_full(ctx)==0:
                        self.aoosat_reg(ctx)
                        if self.aoosun_full(ctx)==0:
                            self.aoosun_reg(ctx)
                            await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SATURDAY 20 UTC** and **SUNDAY 20 UTC** You voted for Sunday.")
                        else:
                            self.sat_addvote(ctx)
                            self.sun_deletevote(ctx)
                            await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SATURDAY 20 UTC**. Sunday is already full. Your vote changed to Saturday.")
                    else:
                        if self.aoosun_full(ctx)==0:
                            self.aoosun_reg(ctx)
                            await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SUNDAY 20 UTC**. Saturday is already full. You voted for Sunday.")
                        else:
                            self.sat_deletevote(ctx)
                            await ctx.send(f"{ctx.author.mention} Sadly both days are full. You didn't vote.")
                #already registered for sun
                elif regsat==0 and regsun==1:
                    if self.aoosat_full(ctx)==0:
                        self.aoosat_reg(ctx)
                        await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SATURDAY 20 UTC**. You were already registered for Sunday. Your vote didn't change.")
                    else: 
                        await ctx.send(f"{ctx.author.mention} Saturday is full already. You were already registered for Sunday. Your vote didn't change.")
                #already registered for sat
                elif regsat==1 and regsun==0:
                    if self.aoosun_full(ctx)==0:
                        self.aoosun_reg(ctx)
                        await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SUNDAY 20 UTC**. You were already registered for Saturday. You voted for Sunday.")
                    else:
                        await ctx.send(f"{ctx.author.mention} Sunday is full already. You were already registered for Saturday. Nothing changed.")
                #already registered for both
                elif regsat==1 and regsun==1:
                    await ctx.send(f" {ctx.author.mention} You are already registered for both days. You voted Sunday")            
            #registering for WgDF
            elif arg=="F" or arg=="f":
                self.aooF(ctx)
                if registered==0:
                    await ctx.send(f"{ctx.author.mention} has been added to the AOO roster for **SUNDAY 4 UTC** at **WgDF**")
                else:
                    await ctx.send(f"{ctx.author.mention} is already registered.")
            else:
                await ctx.send(f"{ctx.author.mention} Please insert valid argument.")
        #if not in registration channel
        else:
            await ctx.send(f"{ctx.author.mention} Please use the registration channel.")